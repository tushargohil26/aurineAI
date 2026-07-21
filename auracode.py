import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from uuid import uuid4

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.syntax import Syntax
    from rich.box import ROUNDED, MINIMAL
except ImportError:
    print("Install rich: pip install rich")
    sys.exit(1)

console = Console()

WORKSPACE = Path.cwd().resolve()

_HAS_AI = False
_HAS_CFG = False
_HAS_Q = False

# ============================================================================
# SESSIONS DIR
# ============================================================================
SESSIONS_DIR = WORKSPACE / ".auracode" / "sessions"
SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
CONFIG_FILE = WORKSPACE / ".auracode" / "config.json"
CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)

# ============================================================================
# AGENTS - OpenCode style sub-agents
# ============================================================================
AGENTS = [
    {"id": "general", "name": "Aurine", "icon": "AI", "desc": "Full-stack AI assistant", "color": "cyan",
     "system": "You are AuraCode, an elite AI assistant. You can read ANY file on the device (Excel, ZIP, PDF, code, data). Give REAL, HELPFUL answers. Never say 'I can't access files outside workspace' - you CAN access any file. Use /open or read_file tool with absolute paths."},
    {"id": "coder", "name": "Coder", "icon": "**", "desc": "Elite coding agent", "color": "green",
     "system": "You are an expert programmer. Write complete, production-ready code. Read any code file on the device. Understand ALL programming languages. Always include imports and error handling."},
    {"id": "researcher", "name": "Researcher", "icon": "🔍", "desc": "Web search & analysis", "color": "blue",
     "system": "You are a research agent. Search the web, analyze information, provide well-sourced answers. Read and analyze any file type."},
    {"id": "planner", "name": "Planner", "icon": "📋", "desc": "Task decomposition", "color": "yellow",
     "system": "You are a project planner. Break complex tasks into clear, actionable steps with estimates."},
    {"id": "creative", "name": "Creative", "icon": "🎨", "desc": "Content generation", "color": "magenta",
     "system": "You are a creative writer. Generate compelling content, marketing copy, documentation."},
    {"id": "data", "name": "Data Analyst", "icon": "📊", "desc": "Data analysis & charts", "color": "red",
     "system": "You are a data analyst. Read Excel, CSV, JSON files from ANY location. Analyze data, create visualizations, write analysis code."},
    {"id": "devops", "name": "DevOps", "icon": "⚙️", "desc": "Deploy & infrastructure", "color": "bright_black",
     "system": "You are a DevOps expert. Handle Docker, CI/CD, deployment, infrastructure."},
    {"id": "tutor", "name": "Tutor", "icon": "📚", "desc": "Learning & teaching", "color": "bright_blue",
     "system": "You are a coding tutor. Explain concepts clearly, provide examples, guide learning. Understand ALL programming languages."},
    {"id": "security", "name": "Security", "icon": "🛡️", "desc": "Vulnerability scanning", "color": "bright_red",
     "system": "You are a security expert. Review code for vulnerabilities, suggest fixes."},
]


def _load_ai():
    global _HAS_AI, _HAS_CFG
    if _HAS_AI:
        return
    try:
        from app.llm import chat_completion, chat_completion_stream, FREE_PROVIDER_URLS
        from app.device import (
            get_device_id, get_user_id, get_device_name,
            store_chat_message, get_chat_history, get_all_chats,
            store_fact, recall_facts, store_preference, get_preferences,
            learn_pattern, get_learned_patterns,
            build_memory_context, extract_facts_from_message,
            get_data_dir, get_chats_dir,
        )
        _HAS_AI = True
    except ImportError:
        pass
    try:
        from app.config import get_settings
        _HAS_CFG = True
    except ImportError:
        pass


# ============================================================================
# MODELS
# ============================================================================
MODELS = [
    {"id": "google/gemini-2.0-flash", "name": "Gemini 2.0 Flash", "provider": "google", "free": True, "tier": "fast", "color": "blue", "key_env": "GOOGLE_API_KEY"},
    {"id": "google/gemini-2.5-pro-preview", "name": "Gemini 2.5 Pro", "provider": "google", "free": True, "tier": "best", "color": "blue", "key_env": "GOOGLE_API_KEY"},
    {"id": "groq/llama-3.3-70b-versatile", "name": "Llama 3.3 70B (Groq)", "provider": "groq", "free": True, "tier": "fast", "color": "bright_magenta", "key_env": "GROQ_API_KEY"},
    {"id": "groq/llama-3.1-8b-instant", "name": "Llama 3.1 8B Instant", "provider": "groq", "free": True, "tier": "fast", "color": "bright_magenta", "key_env": "GROQ_API_KEY"},
    {"id": "groq/mixtral-8x7b-32768", "name": "Mixtral 8x7B", "provider": "groq", "free": True, "tier": "fast", "color": "bright_magenta", "key_env": "GROQ_API_KEY"},
    {"id": "deepseek/deepseek-chat", "name": "DeepSeek Chat", "provider": "deepseek", "free": True, "tier": "best", "color": "green", "key_env": "DEEPSEEK_API_KEY"},
    {"id": "deepseek/deepseek-coder", "name": "DeepSeek Coder", "provider": "deepseek", "free": True, "tier": "best", "color": "green", "key_env": "DEEPSEEK_API_KEY"},
    {"id": "openrouter/meta-llama/llama-3.1-405b-instruct", "name": "Llama 3.1 405B (OpenRouter)", "provider": "openrouter", "free": True, "tier": "best", "color": "bright_yellow", "key_env": "OPENROUTER_API_KEY"},
    {"id": "openrouter/meta-llama/llama-3.1-70b-instruct", "name": "Llama 3.1 70B (OpenRouter)", "provider": "openrouter", "free": True, "tier": "balanced", "color": "bright_yellow", "key_env": "OPENROUTER_API_KEY"},
    {"id": "openrouter/google/gemini-2.0-flash:free", "name": "Gemini Flash (OpenRouter)", "provider": "openrouter", "free": True, "tier": "fast", "color": "bright_yellow", "key_env": "OPENROUTER_API_KEY"},
    {"id": "sambanova/Meta-Llama-3.1-405B-Instruct", "name": "Llama 3.1 405B (SambaNova)", "provider": "sambanova", "free": True, "tier": "best", "color": "bright_green", "key_env": "SAMBANOVA_API_KEY"},
    {"id": "cerebras/llama-3.3-70b", "name": "Llama 3.3 70B (Cerebras)", "provider": "cerebras", "free": True, "tier": "fast", "color": "bright_cyan", "key_env": "CEREBRAS_API_KEY"},
    {"id": "nvidia/meta/llama-3.1-405b-instruct", "name": "Llama 3.1 405B (NVIDIA)", "provider": "nvidia", "free": True, "tier": "best", "color": "bright_red", "key_env": "NVIDIA_API_KEY"},
    {"id": "openai/gpt-4o-mini", "name": "GPT-4o Mini", "provider": "openai", "free": False, "tier": "fast", "color": "bright_green", "key_env": "OPENAI_API_KEY"},
    {"id": "openai/gpt-4o", "name": "GPT-4o", "provider": "openai", "free": False, "tier": "best", "color": "bright_green", "key_env": "OPENAI_API_KEY"},
]

# ============================================================================
# PROVIDERS - for /connect
# ============================================================================
PROVIDERS = [
    {"id": "google", "name": "Google Gemini", "key_env": "GOOGLE_API_KEY", "model_env": "GOOGLE_CHAT_MODEL", "default_model": "gemini-2.0-flash", "url": "https://aistudio.google.com/app/apikey", "free": True, "icon": "G", "color": "blue"},
    {"id": "groq", "name": "Groq", "key_env": "GROQ_API_KEY", "model_env": "GROQ_CHAT_MODEL", "default_model": "llama-3.3-70b-versatile", "url": "https://console.groq.com/keys", "free": True, "icon": "Q", "color": "bright_magenta"},
    {"id": "openai", "name": "OpenAI", "key_env": "OPENAI_API_KEY", "model_env": "OPENAI_CHAT_MODEL", "default_model": "gpt-4o-mini", "url": "https://platform.openai.com/api-keys", "free": False, "icon": "O", "color": "bright_green"},
    {"id": "anthropic", "name": "Anthropic", "key_env": "ANTHROPIC_API_KEY", "model_env": "ANTHROPIC_CHAT_MODEL", "default_model": "claude-sonnet-4-20250514", "url": "https://console.anthropic.com/", "free": False, "icon": "A", "color": "yellow"},
    {"id": "deepseek", "name": "DeepSeek", "key_env": "DEEPSEEK_API_KEY", "model_env": "DEEPSEEK_CHAT_MODEL", "default_model": "deepseek-chat", "url": "https://platform.deepseek.com/", "free": True, "icon": "D", "color": "green"},
]

# ============================================================================
# PLUGINS - Real functionality like OpenCode
# ============================================================================
PLUGINS = [
    {"id": "git", "name": "Git", "icon": "G", "desc": "Version control",
     "actions": ["status", "log", "diff", "commit", "push", "pull", "branches"],
     "run": lambda action, args: _plugin_git(action, args)},
    {"id": "files", "name": "File Manager", "icon": "F", "desc": "File operations",
     "actions": ["list", "read", "write", "search", "info"],
     "run": lambda action, args: _plugin_files(action, args)},
    {"id": "code", "name": "Code Runner", "icon": "X", "desc": "Execute code",
     "actions": ["python", "javascript", "shell"],
     "run": lambda action, args: _plugin_code(action, args)},
    {"id": "data", "name": "Data Analyzer", "icon": "D", "desc": "Analyze any data file",
     "actions": ["excel", "csv", "json", "sqlite", "summary"],
     "run": lambda action, args: _plugin_data(action, args)},
    {"id": "compress", "name": "Compressor", "icon": "Z", "desc": "ZIP/unzip files",
     "actions": ["zip", "unzip", "list"],
     "run": lambda action, args: _plugin_compress(action, args)},
    {"id": "token", "name": "Token Saver", "icon": "$", "desc": "Save tokens by summarizing",
     "actions": ["summarize", "compress", "extract"],
     "run": lambda action, args: _plugin_token(action, args)},
]


def _plugin_git(action, args):
    import subprocess
    cmd = f"git {action} {args}".strip()
    r = subprocess.run(cmd, shell=True, text=True, capture_output=True, timeout=30)
    return (r.stdout + r.stderr).strip()[:5000] or f"git {action} done"


def _plugin_files(action, args):
    if action == "list":
        return _list_files(args or ".")
    elif action == "read":
        return _read_file(args)
    elif action == "info":
        p = Path(args).resolve()
        if p.exists():
            info = f"Name: {p.name}\nSize: {p.stat().st_size} bytes\nType: {p.suffix}\nModified: {datetime.fromtimestamp(p.stat().st_mtime)}"
            return info
        return "File not found"
    return f"Unknown file action: {action}"


def _plugin_code(action, args):
    import subprocess
    if action == "python":
        r = subprocess.run(f'python -c "{args}"', shell=True, text=True, capture_output=True, timeout=30)
    elif action == "javascript":
        r = subprocess.run(f'node -e "{args}"', shell=True, text=True, capture_output=True, timeout=30)
    else:
        r = subprocess.run(args, shell=True, text=True, capture_output=True, timeout=30)
    return (r.stdout + r.stderr).strip()[:5000] or "Done"


def _plugin_data(action, args):
    if action == "excel":
        return _read_excel(Path(args).resolve())
    elif action == "csv":
        return _read_csv(Path(args).resolve())
    elif action == "json":
        return _read_json(Path(args).resolve())
    elif action == "sqlite":
        return _read_sqlite_db(Path(args).resolve())
    elif action == "summary":
        p = Path(args).resolve()
        suffix = p.suffix.lower()
        if suffix == ".xlsx":
            return _read_excel(p)
        elif suffix == ".csv":
            return _read_csv(p)
        elif suffix == ".json":
            return _read_json(p)
        elif suffix in (".db", ".sqlite", ".sqlite3"):
            return _read_sqlite_db(p)
        return _read_file(str(p))
    return f"Unknown data action: {action}"


def _plugin_compress(action, args):
    import zipfile
    import subprocess
    if action == "zip":
        parts = args.split()
        if len(parts) >= 2:
            output = parts[0]
            inputs = parts[1:]
            with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
                for inp in inputs:
                    p = Path(inp)
                    if p.is_dir():
                        for f in p.rglob("*"):
                            if f.is_file():
                                zf.write(f, f.relative_to(p.parent))
                    elif p.is_file():
                        zf.write(p, p.name)
            return f"Created {output} ({Path(output).stat().st_size // 1024}KB)"
    elif action == "unzip":
        parts = args.split()
        if len(parts) >= 1:
            src = parts[0]
            dst = parts[1] if len(parts) > 1 else "."
            with zipfile.ZipFile(src, 'r') as zf:
                zf.extractall(dst)
            return f"Extracted to {dst}"
    elif action == "list":
        with zipfile.ZipFile(args, 'r') as zf:
            return "\n".join(zf.namelist()[:50])
    return f"Unknown compress action: {action}"


def _plugin_token(action, args):
    if action == "summarize":
        words = args.split()
        if len(words) > 100:
            return " ".join(words[:100]) + f"... ({len(words)} words total)"
        return args
    elif action == "compress":
        lines = args.split("\n")
        if len(lines) > 50:
            return "\n".join(lines[:50]) + f"\n... ({len(lines)} lines total)"
        return args
    elif action == "extract":
        import re
        urls = re.findall(r'https?://\S+', args)
        emails = re.findall(r'\S+@\S+', args)
        return f"URLs: {len(urls)}\nEmails: {len(emails)}\n" + "\n".join(urls[:10])
    return args

# ============================================================================
# COMMANDS - OpenCode style
# ============================================================================
COMMANDS = [
    {"cmd": "/help", "icon": "?", "desc": "Show all commands", "group": "General", "key": "Ctrl+H"},
    {"cmd": "/connect", "icon": "C", "desc": "Connect AI provider", "group": "General", "key": "Ctrl+K"},
    {"cmd": "/agents", "icon": "A", "desc": "Select AI agent", "group": "General", "key": "Ctrl+A"},
    {"cmd": "/model", "icon": "M", "desc": "Select AI model", "group": "General", "key": "Ctrl+M"},
    {"cmd": "/doctor", "icon": "D", "desc": "System diagnostics", "group": "General", "key": ""},
    {"cmd": "/config", "icon": "X", "desc": "Configuration", "group": "General", "key": ""},
    {"cmd": "/plugins", "icon": "P", "desc": "Manage plugins", "group": "General", "key": "Ctrl+P"},
    {"cmd": "/skills", "icon": "S", "desc": "Available skills", "group": "General", "key": ""},
    {"cmd": "/session", "icon": "B", "desc": "Switch session", "group": "Sessions", "key": "Ctrl+S"},
    {"cmd": "/sessions", "icon": "Z", "desc": "List all sessions", "group": "Sessions", "key": ""},
    {"cmd": "/new", "icon": "N", "desc": "New session", "group": "Sessions", "key": "Ctrl+N"},
    {"cmd": "/resume", "icon": "R", "desc": "Resume last session", "group": "Sessions", "key": ""},
    {"cmd": "/files", "icon": "F", "desc": "List files", "group": "Files", "key": ""},
    {"cmd": "/read", "icon": "R", "desc": "Read file", "group": "Files", "key": ""},
    {"cmd": "/open", "icon": "O", "desc": "Open ANY file (Excel, ZIP, PDF, etc)", "group": "Files", "key": ""},
    {"cmd": "/write", "icon": "W", "desc": "Write file", "group": "Files", "key": ""},
    {"cmd": "/edit", "icon": "E", "desc": "Edit file", "group": "Files", "key": ""},
    {"cmd": "/run", "icon": "!", "desc": "Run command", "group": "Files", "key": ""},
    {"cmd": "/diff", "icon": "D", "desc": "Git diff", "group": "Git", "key": ""},
    {"cmd": "/log", "icon": "L", "desc": "Git log", "group": "Git", "key": ""},
    {"cmd": "/git", "icon": "G", "desc": "Git command", "group": "Git", "key": ""},
    {"cmd": "/commit", "icon": "C", "desc": "Git commit", "group": "Git", "key": ""},
    {"cmd": "/init", "icon": "I", "desc": "Init workspace", "group": "Tools", "key": ""},
    {"cmd": "/web", "icon": "W", "desc": "Search web", "group": "Tools", "key": ""},
    {"cmd": "/img", "icon": "I", "desc": "Analyze image", "group": "Tools", "key": ""},
    {"cmd": "/history", "icon": "H", "desc": "Chat history", "group": "Tools", "key": ""},
    {"cmd": "/device", "icon": "V", "desc": "Device info", "group": "Tools", "key": ""},
    {"cmd": "/compact", "icon": "Q", "desc": "Compress context", "group": "Tools", "key": ""},
    {"cmd": "/cost", "icon": "$", "desc": "Token estimate", "group": "Tools", "key": ""},
    {"cmd": "/clear", "icon": "X", "desc": "Clear screen", "group": "General", "key": "Ctrl+L"},
    {"cmd": "/quit", "icon": "Q", "desc": "Exit", "group": "General", "key": "Ctrl+Q"},
]

# ============================================================================
# SYSTEM PROMPT
# ============================================================================
SYSTEM_PROMPT = """You are AuraCode, an AI that RUNS commands and READS files. You are NOT a chatbot - you EXECUTE actions.

IMPORTANT: When user asks something, you MUST return actions to do it. NEVER just give advice.

EXAMPLES:
User: "git status" -> {"message": "Running git status...", "actions": [{"tool": "run_command", "command": "git status"}]}
User: "what files are here" -> {"message": "Listing files...", "actions": [{"tool": "list_files", "path": "."}]}
User: "read this file C:\\test.py" -> {"message": "Reading file...", "actions": [{"tool": "read_file", "path": "C:\\test.py"}]}
User: "connect to git" -> {"message": "Checking git status...", "actions": [{"tool": "run_command", "command": "git status"}, {"tool": "run_command", "command": "git remote -v"}]}
User: "show me my repos" -> {"message": "Listing git repositories...", "actions": [{"tool": "run_command", "command": "dir /b"}]}
User: "create a file" -> {"message": "Creating file...", "actions": [{"tool": "write_file", "path": "newfile.txt", "content": "content here"}]}
User: "run python code" -> {"message": "Running code...", "actions": [{"tool": "run_command", "command": "python -c 'print(1+1)'"}]}
User: "analyze this excel" -> {"message": "Reading Excel...", "actions": [{"tool": "read_file", "path": "path_to_file"}]}

TOOLS YOU MUST USE:
- run_command: Run any shell command (git, python, dir, etc)
- read_file: Read any file (Excel, PDF, code, text)
- write_file: Create or write files
- list_files: List directory contents

RULES:
1. ALWAYS return actions - never just give text advice
2. When user says "git" -> run git commands immediately
3. When user says "connect" -> run connection commands
4. When user gives a file path -> read it immediately
5. When user asks "can you" -> DO IT, don't just say yes
6. Return JSON: {"message": "what you're doing", "actions": [{"tool": "...", ...}]}
"""

# ============================================================================
# STATE
# ============================================================================
_state = {
    "chat_id": None,
    "agent": AGENTS[0],
    "model": MODELS[0],
    "history": [],
    "sessions": {},
    "session_name": None,
    "provider": "aurine",
    "model_name": "gpt-4o-mini",
}

# ============================================================================
# HELPERS
# ============================================================================
IGNORE_DIRS = {".git", ".venv", "__pycache__", "node_modules", ".mypy_cache", ".pytest_cache", ".auracode"}


def _safe(p):
    c = Path(p).resolve()
    return c


def _list_files(path="."):
    root = Path(path).resolve() if Path(path).is_absolute() else (WORKSPACE / path).resolve()
    if not root.exists():
        return "Path does not exist."
    lines = []

    def _walk(d, pre="", dep=0):
        if dep > 4:
            return
        try:
            ents = sorted(d.iterdir(), key=lambda e: (not e.is_dir(), e.name.lower()))
        except PermissionError:
            lines.append(f"{pre}[red]Permission denied[/]")
            return
        ents = [e for e in ents if e.name not in IGNORE_DIRS and e.suffix != ".pyc"]
        for i, e in enumerate(ents):
            last = i == len(ents) - 1
            conn = "└─── " if last else "├─── "
            if e.is_dir():
                lines.append(f"{pre}{conn}[blue]{e.name}/[/]")
                _walk(e, pre + ("    " if last else "│   "), dep + 1)
            else:
                sz = e.stat().st_size
                s = f"{sz // 1048576}MB" if sz > 1048576 else f"{sz // 1024}KB" if sz > 1024 else f"{sz}B"
                lines.append(f"{pre}{conn}{e.name}  [dim]({s})[/]")

    _walk(root)
    return "\n".join(lines[:120]) or "No files."


def _read_file(p):
    t = Path(p).resolve() if Path(p).is_absolute() else (WORKSPACE / p).resolve()
    if not t.exists() or not t.is_file():
        return "File not found."
    suffix = t.suffix.lower()
    if suffix == ".zip":
        return _read_zip(t)
    if suffix in (".xlsx", ".xls"):
        return _read_excel(t)
    if suffix == ".pdf":
        return _read_pdf(t)
    if suffix in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
        return f"[Image file: {t.name} ({t.stat().st_size // 1024}KB). Use /img to analyze visually.]"
    if suffix in (".mp3", ".wav", ".mp4", ".avi", ".mkv"):
        return f"[Media file: {t.name} ({t.stat().st_size // 1024}KB)]"
    if suffix in (".db", ".sqlite", ".sqlite3"):
        return _read_sqlite_db(t)
    if suffix in (".json",):
        return _read_json(t)
    if suffix in (".csv",):
        return _read_csv(t)
    try:
        return t.read_text(encoding="utf-8", errors="ignore")[:30000]
    except Exception as e:
        return f"Cannot read {t.name}: {e}"


def _read_zip(p):
    import zipfile
    try:
        with zipfile.ZipFile(str(p), 'r') as z:
            names = z.namelist()
            total = len(names)
            dirs = sum(1 for n in names if n.endswith('/'))
            files = total - dirs
            result = f"ZIP: {p.name} ({p.stat().st_size // 1024}KB)\n"
            result += f"Contains: {files} files, {dirs} directories\n\n"
            result += "Contents:\n"
            for name in names[:50]:
                info = z.getinfo(name)
                sz = info.file_size
                s = f"{sz // 1048576}MB" if sz > 1048576 else f"{sz // 1024}KB" if sz > 1024 else f"{sz}B"
                if name.endswith('/'):
                    result += f"  {name}  [dir]\n"
                else:
                    result += f"  {name}  ({s})\n"
            if total > 50:
                result += f"  ... and {total - 50} more files\n"
            return result
    except Exception as e:
        return f"Cannot read ZIP: {e}"


def _read_excel(p):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        result = f"Excel: {p.name}\n"
        result += f"Sheets: {', '.join(wb.sheetnames)}\n\n"
        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            result += f"--- Sheet: {sheet_name} ---\n"
            rows = list(ws.iter_rows(max_row=20, values_only=True))
            if rows:
                for i, row in enumerate(rows[:20]):
                    vals = [str(v) if v is not None else "" for v in row]
                    result += " | ".join(vals) + "\n"
                    if i == 0:
                        result += "-" * 60 + "\n"
            result += "\n"
        wb.close()
        return result[:15000]
    except ImportError:
        return f"Excel file: {p.name}. Install openpyxl for Excel support: pip install openpyxl"
    except Exception as e:
        return f"Cannot read Excel: {e}"


def _read_pdf(p):
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(p))
        result = f"PDF: {p.name} ({len(reader.pages)} pages)\n\n"
        for i, page in enumerate(reader.pages[:10]):
            text = page.extract_text()
            if text:
                result += f"--- Page {i+1} ---\n{text[:2000]}\n\n"
        return result[:15000]
    except ImportError:
        return f"PDF file: {p.name}. Install pypdf for PDF support: pip install pypdf"
    except Exception as e:
        return f"Cannot read PDF: {e}"


def _read_json(p):
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
        result = f"JSON: {p.name}\n\n"
        result += json.dumps(data, indent=2, ensure_ascii=False)[:15000]
        return result
    except Exception as e:
        return f"Cannot read JSON: {e}"


def _read_csv(p):
    try:
        import csv
        result = f"CSV: {p.name}\n\n"
        with open(str(p), 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 25:
                    result += f"... ({sum(1 for _ in open(str(p)))} total rows)\n"
                    break
                result += " | ".join(row) + "\n"
                if i == 0:
                    result += "-" * 60 + "\n"
        return result[:15000]
    except Exception as e:
        return f"Cannot read CSV: {e}"


def _read_sqlite_db(p):
    try:
        import sqlite3
        conn = sqlite3.connect(str(p))
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [row[0] for row in cursor.fetchall()]
        result = f"SQLite DB: {p.name}\nTables: {', '.join(tables)}\n\n"
        for table in tables[:10]:
            cursor.execute(f"SELECT COUNT(*) FROM [{table}]")
            count = cursor.fetchone()[0]
            result += f"--- {table} ({count} rows) ---\n"
            cursor.execute(f"SELECT * FROM [{table}] LIMIT 5")
            cols = [desc[0] for desc in cursor.description]
            result += " | ".join(cols) + "\n"
            result += "-" * 60 + "\n"
            for row in cursor.fetchall():
                result += " | ".join(str(v) for v in row) + "\n"
            result += "\n"
        conn.close()
        return result[:15000]
    except Exception as e:
        return f"Cannot read SQLite: {e}"


def _write_file(p, c):
    t = Path(p).resolve() if Path(p).is_absolute() else (WORKSPACE / p).resolve()
    t.parent.mkdir(parents=True, exist_ok=True)
    t.write_text(c, encoding="utf-8")
    return f"Wrote {t}"


def _edit_file(p, old, new):
    t = Path(p).resolve() if Path(p).is_absolute() else (WORKSPACE / p).resolve()
    if not t.exists():
        return "File not found."
    content = t.read_text(encoding="utf-8")
    if old not in content:
        return "Old text not found in file."
    count = content.count(old)
    if count > 1:
        return f"Ambiguous: {count} matches found. Provide more context."
    new_content = content.replace(old, new, 1)
    t.write_text(new_content, encoding="utf-8")
    return f"Edited {t}"


def _run_cmd(cmd):
    bl = ["rm ", "del ", "format ", "git reset", "rmdir ", "Remove-Item"]
    if any(x.lower() in cmd.lower() for x in bl):
        return "Blocked destructive command."
    r = subprocess.run(cmd, cwd=WORKSPACE, shell=True, text=True, capture_output=True, timeout=60)
    return (r.stdout + r.stderr).strip()[:30000] or f"Exit code {r.returncode}."


def _exec(actions):
    res = []
    for a in actions:
        t = a.get("tool")
        try:
            if t == "list_files":
                r = _list_files(a.get("path", "."))
            elif t == "read_file":
                r = _read_file(a.get("path", ""))
            elif t == "write_file":
                r = _write_file(a.get("path", ""), a.get("content", ""))
            elif t == "run_command":
                r = _run_cmd(a.get("command", ""))
            else:
                r = f"Unknown tool: {t}"
        except Exception as e:
            r = f"Error: {e}"
        res.append(f"  [cyan]▶ {t}[/] [dim]{r[:500]}[/]")
    return "\n".join(res)


def _get_provider_info():
    from app.config import _read_env
    _read_env()
    provider = os.getenv("AI_PROVIDER", "google").strip().lower()

    model = os.getenv("GOOGLE_CHAT_MODEL") or os.getenv("OPENAI_CHAT_MODEL") or os.getenv("GROQ_CHAT_MODEL") or "gemini-2.0-flash"
    has_key = bool(
        os.getenv("GOOGLE_API_KEY", "").strip() or
        os.getenv("OPENAI_API_KEY", "").strip() or
        os.getenv("GROQ_API_KEY", "").strip() or
        os.getenv("DEEPSEEK_API_KEY", "").strip() or
        os.getenv("ANTHROPIC_API_KEY", "").strip()
    )
    return provider, model, has_key


def _get_device_str():
    if _HAS_AI:
        return f"{get_device_name()} [dim]id:{get_device_id()[:12]}[/]"
    import platform
    return f"{platform.node()} ({platform.system()})"


def _load_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text())
        except Exception:
            pass
    return {}


def _save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2))


def _fuzzy_match(query, text):
    query = query.lower()
    text_lower = text.lower()
    if query in text_lower:
        return 1.0
    matches = difflib.SequenceMatcher(None, query, text_lower).ratio()
    return matches


def _fuzzy_filter(query, items, key_func):
    if not query:
        return items
    scored = []
    for item in items:
        score = _fuzzy_match(query, key_func(item))
        if score > 0.3:
            scored.append((score, item))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [item for _, item in scored]


# ============================================================================
# SESSION MANAGEMENT
# ============================================================================

def _get_sessions():
    sessions = []
    for f in SESSIONS_DIR.glob("*.json"):
        try:
            data = json.loads(f.read_text())
            data["_file"] = str(f)
            sessions.append(data)
        except Exception:
            pass
    sessions.sort(key=lambda s: s.get("updated", ""), reverse=True)
    return sessions


def _save_session(session_id, messages, name=""):
    path = SESSIONS_DIR / f"{session_id}.json"
    data = {
        "id": session_id,
        "name": name or session_id[:8],
        "messages": messages,
        "created": datetime.now().isoformat(),
        "updated": datetime.now().isoformat(),
        "agent": _state["agent"]["id"],
        "model": _state["model"]["id"],
        "workspace": str(WORKSPACE),
    }
    path.write_text(json.dumps(data, indent=2, default=str))


def _load_session(session_id):
    path = SESSIONS_DIR / f"{session_id}.json"
    if path.exists():
        return json.loads(path.read_text())
    return None


def _list_sessions():
    sessions = _get_sessions()
    if not sessions:
        return None
    table = Table(box=ROUNDED, title="[bold cyan]Sessions[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
    table.add_column("ID", style="dim", min_width=12)
    table.add_column("Name", style="bold")
    table.add_column("Messages", justify="right")
    table.add_column("Agent", style="cyan")
    table.add_column("Updated", style="dim")
    for s in sessions:
        msg_count = len(s.get("messages", []))
        updated = s.get("updated", "")[:16]
        table.add_row(s.get("id", "")[:12], s.get("name", ""), str(msg_count), s.get("agent", "general"), updated)
    return table


# ============================================================================
# Fuzzy Selection (like OpenCode Ctrl+P)
# ============================================================================

def _fuzzy_select(title, options, key_func=None, multi=False):
    """Interactive fuzzy-search selection like OpenCode's Ctrl+P command palette."""
    if not _HAS_Q:
        if multi:
            return []
        return None

    if not options:
        return [] if multi else None

    def _get_display(opt):
        if isinstance(opt, dict):
            return opt.get("display", opt.get("name", opt.get("cmd", str(opt))))
        return str(opt)

    def _get_value(opt):
        if isinstance(opt, dict):
            return opt.get("value", opt)
        return opt

    query = ""
    filtered = list(options)

    while True:
        display_items = []
        for opt in filtered[:20]:
            d = _get_display(opt)
            display_items.append(questionary.Choice(title=d, value=opt))

        if not display_items:
            console.print("[dim]No matches found. Press Esc to go back.[/]")
            return [] if multi else None

        result = questionary.select(
            title,
            choices=display_items,
            style=questionary.Style([
                ('question', 'bold cyan'),
                ('pointer', 'fg:cyan bold'),
                ('highlighted', 'fg:cyan bold'),
                ('selected', 'fg:green bold'),
            ]),
        ).ask()

        if result is None:
            return [] if multi else None
        return _get_value(result)


def _command_palette():
    """Ctrl+P style command palette - the main entry point for all commands."""
    console.print()
    console.print("[bold cyan]Command Palette[/]  [dim]- Type to filter, Enter to select, Esc to cancel[/]")
    console.print()

    options = []
    for cmd in COMMANDS:
        key_hint = f"  [dim]{cmd['key']}[/]" if cmd.get("key") else ""
        options.append({
            "display": f"{cmd['icon']}  {cmd['cmd']}  [dim]{cmd['desc']}[/]{key_hint}",
            "value": cmd["cmd"],
            "cmd": cmd["cmd"],
        })

    result = _fuzzy_select("🔍 Search commands...", options, key_func=lambda o: o.get("cmd", ""))
    if result:
        console.print()
        return result
    return None


# ============================================================================
# /connect - Provider Connection
# ============================================================================

def _show_connect():
    """Interactive provider connection."""
    console.print()
    console.print(Panel("[bold cyan]Connect to AI Provider[/]\n[dim]Get free key: https://aistudio.google.com/app/apikey[/]", border_style="cyan", padding=(0, 1)))

    status_table = Table(box=MINIMAL, border_style="cyan", padding=(0, 1))
    status_table.add_column("Provider", style="bold")
    status_table.add_column("Status", min_width=12)
    status_table.add_column("Key", style="dim")
    status_table.add_column("Free", min_width=5)

    for p in PROVIDERS:
        key_val = os.getenv(p["key_env"], "") if p["key_env"] else ""
        is_set = bool(key_val)
        status = "[green]✓ connected[/]" if is_set else "[red]✗ not set[/]"
        key_display = f"{key_val[:8]}..." if len(key_val) > 8 else (key_val if key_val else "—")
        free = "[green]✓[/]" if p["free"] else "[red]✗[/]"
        status_table.add_row(f"{p['icon']} {p['name']}", status, key_display, free)

    console.print(status_table)
    console.print()

    options = []
    for p in PROVIDERS:
        key_val = os.getenv(p["key_env"], "") if p["key_env"] else ""
        is_set = bool(key_val)
        status = "[green]✓[/]" if is_set else "[red]✗[/]"
        free = " [green](free)[/]" if p["free"] else " [red](paid)[/]"
        options.append({
            "display": f"{p['icon']}  {p['name']}  {status}{free}  [dim]{p['url']}[/]",
            "value": p,
        })

    result = _fuzzy_select("Select provider to configure:", options, key_func=lambda o: o.get("name", ""))
    if not result:
        return

    provider = result
    console.print()

    console.print(f"[bold cyan]Connecting to {provider['name']}[/]")
    console.print(f"[dim]Get your API key: {provider['url']}[/]")
    console.print()

    current_key = os.getenv(provider["key_env"], "") if provider["key_env"] else ""
    hint = f"  [dim](current: {current_key[:8]}...)[/]" if current_key else ""
    console.print(f"[dim]Paste your {provider['name']} API key:{hint}[/]")

    try:
        api_key = input("  API Key: ").strip()
    except (EOFError, KeyboardInterrupt):
        console.print("\n[dim]Cancelled[/]")
        return

    if not api_key and current_key:
        console.print("[dim]Keeping existing key[/]")
    elif api_key:
        _set_env_var(provider["key_env"], api_key)
        console.print(f"\n[green]✓[/] API key saved for {provider['name']}")
    else:
        console.print("[yellow]No key entered[/]")
        return

    _set_env_var("AI_PROVIDER", provider["id"])
    if provider.get("model_env"):
        _set_env_var(provider["model_env"], provider["default_model"])

    console.print(f"[green]✓[/] Provider: [bold]{provider['name']}[/]")
    console.print(f"[green]✓[/] Model: [bold]{provider['default_model']}[/]")
    console.print(f"\n[dim]Restart AuraCode to apply changes, or type /model to switch models.[/]\n")


def _set_env_var(key, value):
    """Set an environment variable in .env file."""
    env_path = WORKSPACE / ".env"
    lines = []
    found = False
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.strip().startswith(f"{key}="):
                lines.append(f"{key}={value}")
                found = True
            else:
                lines.append(line)
    if not found:
        lines.append(f"{key}={value}")
    env_path.write_text("\n".join(lines) + "\n")
    os.environ[key] = value
    # Reload settings so the new key takes effect immediately
    try:
        from app.config import reload_settings
        reload_settings()
    except ImportError:
        pass


# ============================================================================
# RICH DISPLAY FUNCTIONS
# ============================================================================

def _show_help():
    groups = {}
    for cmd in COMMANDS:
        g = cmd["group"]
        if g not in groups:
            groups[g] = []
        groups[g].append(cmd)

    for group_name, cmds in groups.items():
        table = Table(box=ROUNDED, title=f"[bold cyan]{group_name}[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1), min_width=50)
        table.add_column("Command", style="bold yellow", min_width=14)
        table.add_column("Description", style="white")
        table.add_column("Key", style="dim")
        for cmd in cmds:
            table.add_row(cmd["cmd"], cmd["desc"], cmd.get("key", ""))
        console.print()
        console.print(table)
    console.print()


def _show_agents():
    table = Table(box=ROUNDED, title="[bold cyan]AI Agents[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
    table.add_column("", min_width=3)
    table.add_column("Name", style="bold", min_width=16)
    table.add_column("Description", style="white")
    table.add_column("ID", style="dim")
    for a in AGENTS:
        marker = " [green]◄ active[/]" if a["id"] == _state["agent"]["id"] else ""
        table.add_row(a["icon"], f"[{a['color']}]{a['name']}[/]{marker}", a["desc"], a["id"])
    console.print()
    console.print(table)
    console.print()


def _show_models():
    from app.config import _read_env
    _read_env()
    table = Table(box=ROUNDED, title="[bold cyan]AI Models[/] [dim]- All latest, high quality[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
    table.add_column("Model", style="bold", min_width=32)
    table.add_column("Provider", style="cyan", min_width=12)
    table.add_column("Tier", min_width=8)
    table.add_column("Free", min_width=5)
    table.add_column("Key", min_width=8)
    for m in MODELS:
        tier_color = "green" if m["tier"] == "fast" else "yellow" if m["tier"] == "balanced" else "red"
        free_mark = "[green]✓[/]" if m["free"] else "[red]✗[/]"
        key_set = bool(os.getenv(m.get("key_env", ""), "")) if m.get("key_env") else False
        key_mark = "[green]✓[/]" if key_set else "[dim]—[/]"
        is_active = m["id"] == _state["model"]["id"]
        marker = " [green]◄[/]" if is_active else ""
        table.add_row(f"[{m['color']}]{m['name']}[/]{marker}", m["provider"], f"[{tier_color}]{m['tier']}[/]", free_mark, key_mark)
    console.print()
    console.print(table)
    console.print()


def _show_plugins():
    table = Table(box=ROUNDED, title="[bold cyan]Plugins[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
    table.add_column("", min_width=3)
    table.add_column("Name", style="bold", min_width=14)
    table.add_column("Description", style="white")
    table.add_column("Actions", style="dim")
    for p in PLUGINS:
        actions = ", ".join(p["actions"][:5])
        table.add_row(p["icon"], f"[cyan]{p['name']}[/]", p["desc"], actions)
    console.print()
    console.print(table)
    console.print()


def _show_skills():
    skills = [
        ("Code Review", "Review code for bugs and improvements"),
        ("Refactor", "Refactor code for better structure"),
        ("Test Generator", "Generate unit tests"),
        ("Doc Generator", "Generate documentation"),
        ("API Design", "Design REST/GraphQL APIs"),
        ("Database Design", "Schema design and queries"),
        ("Security Audit", "Scan for vulnerabilities"),
        ("Performance", "Optimize code performance"),
        ("Git Workflow", "Branch strategy and commits"),
        ("Docker", "Dockerfile and compose setup"),
        ("CI/CD", "GitHub Actions, pipelines"),
        ("Deploy", "Deployment automation"),
    ]
    table = Table(box=ROUNDED, title="[bold cyan]Skills[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
    table.add_column("Skill", style="bold green", min_width=18)
    table.add_column("Description", style="white")
    for name, desc in skills:
        table.add_row(name, desc)
    console.print()
    console.print(table)
    console.print()


def _show_init():
    files = ["package.json", "requirements.txt", "Cargo.toml", "go.mod", "pom.xml", "package-lock.json", "pyproject.toml"]
    found = [f for f in files if (WORKSPACE / f).exists()]
    panel_content = []
    if found:
        panel_content.append(f"[green]✓[/] Project: {', '.join(found)}")
    else:
        panel_content.append(f"[yellow]![/] No standard project files in [bold]{WORKSPACE.name}[/]")
    dotgit = WORKSPACE / ".git"
    if dotgit.exists():
        branch = _run_cmd("git branch --show-current").strip()
        remote = _run_cmd("git remote get-url origin").strip()
        panel_content.append(f"[green]✓[/] Git: branch={branch or 'none'} remote={remote or 'none'}")
    else:
        panel_content.append(f"[yellow]![/] No git repo")
    panel_content.append(f"[green]✓[/] Workspace: [bold]{WORKSPACE}[/]")
    if _HAS_AI:
        panel_content.append(f"[green]✓[/] Device: {_get_device_str()}")
    prov, model, has_key = _get_provider_info()
    status = "[green]✓[/]" if has_key else "[yellow]![/]"
    panel_content.append(f"{status} Provider: [bold]{prov}[/]  Model: {model}")
    console.print()
    console.print(Panel("\n".join(panel_content), title="[bold cyan]Workspace Init[/]", border_style="cyan"))
    console.print()


def _show_config():
    from app.config import _read_env
    _read_env()
    table = Table(box=ROUNDED, title="[bold cyan]Configuration[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
    table.add_column("Key", style="bold yellow", min_width=24)
    table.add_column("Status", min_width=12)
    table.add_column("Value", style="dim")
    keys = {
        "AI_PROVIDER": os.getenv("AI_PROVIDER", "google"),
        "OPENAI_API_KEY": "set" if os.getenv("OPENAI_API_KEY") else "not set",
        "GOOGLE_API_KEY": "set" if os.getenv("GOOGLE_API_KEY") else "not set",
        "GROQ_API_KEY": "set" if os.getenv("GROQ_API_KEY") else "not set",
        "DEEPSEEK_API_KEY": "set" if os.getenv("DEEPSEEK_API_KEY") else "not set",
        "ANTHROPIC_API_KEY": "set" if os.getenv("ANTHROPIC_API_KEY") else "not set",
        "OPENAI_CHAT_MODEL": os.getenv("OPENAI_CHAT_MODEL", "gpt-4o-mini"),
        "GOOGLE_CHAT_MODEL": os.getenv("GOOGLE_CHAT_MODEL", "gemini-2.0-flash"),
        "GROQ_CHAT_MODEL": os.getenv("GROQ_CHAT_MODEL", "llama-3.3-70b-versatile"),
    }
    for k, v in keys.items():
        st = "[green]✓ set[/]" if v == "set" or v == "built-in" else "[red]✗ not set[/]" if v == "not set" else "[dim]"
        table.add_row(k, st, v)
    console.print()
    console.print(table)
    console.print("[dim]Edit .env file or use /connect to change settings[/]")
    console.print()


def _show_doctor():
    import platform
    table1 = Table(box=MINIMAL, border_style="cyan", padding=(0, 1))
    table1.add_column("Check", style="bold cyan", min_width=16)
    table1.add_column("Status", min_width=20)
    table1.add_column("Detail", style="dim")
    table1.add_row("OS", "[green]OK[/]", f"{platform.system()} {platform.release()}")
    table1.add_row("Python", "[green]OK[/]", sys.version.split()[0])
    table1.add_row("Workspace", "[green]OK[/]", str(WORKSPACE))
    if _HAS_AI:
        table1.add_row("Device", "[green]OK[/]", _get_device_str())
    console.print()
    console.print(Panel(table1, title="[bold cyan]System[/]", border_style="cyan"))
    table2 = Table(box=MINIMAL, border_style="cyan", padding=(0, 1))
    table2.add_column("Check", style="bold cyan", min_width=16)
    table2.add_column("Status", min_width=20)
    prov, model, has_key = _get_provider_info()
    table2.add_row("Provider", f"[bold]{prov}[/]")
    table2.add_row("Model", model)
    table2.add_row("API Key", "[green]✓ configured[/]" if has_key else "[red]✗ MISSING - add GOOGLE_API_KEY to .env[/]")
    console.print(Panel(table2, title="[bold cyan]AI Provider[/]", border_style="cyan"))
    table3 = Table(box=MINIMAL, border_style="cyan", padding=(0, 1))
    table3.add_column("Tool", style="bold", min_width=10)
    table3.add_column("Status", min_width=10)
    for tool in ["git", "python", "node", "npm", "docker", "code", "gh"]:
        found = bool(shutil.which(tool))
        st = "[green]✓[/]" if found else "[red]✗[/]"
        table3.add_row(tool, st)
    console.print(Panel(table3, title="[bold cyan]Tools[/]", border_style="cyan"))
    console.print()


def _show_device():
    table = Table(box=ROUNDED, title="[bold cyan]Device Info[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
    table.add_column("Key", style="bold cyan", min_width=16)
    table.add_column("Value", style="white")
    table.add_row("Workspace", str(WORKSPACE))
    if _HAS_AI:
        table.add_row("Device", get_device_name())
        table.add_row("Device ID", get_device_id()[:16])
        table.add_row("User ID", get_user_id())
        table.add_row("Data Dir", str(get_data_dir()))
    table.add_row("Agent", f"{_state['agent']['icon']} {_state['agent']['name']}")
    table.add_row("Model", f"{_state['model']['name']}")
    table.add_row("Session", _state.get("session_name") or _state.get("chat_id", "none")[:12])
    console.print()
    console.print(table)
    console.print()


def _show_diff():
    result = _run_cmd("git diff --stat")
    full = _run_cmd("git diff")
    console.print()
    console.print(Panel(result or "[dim]No changes[/]", title="[bold cyan]Git Diff[/]", border_style="cyan"))
    if full.strip():
        for line in full.split("\n")[:60]:
            color = "green" if line.startswith("+") else "red" if line.startswith("-") else "dim"
            console.print(f"  [{color}]{line}[/]")
    console.print()


def _show_log():
    result = _run_cmd("git log --oneline -20")
    console.print()
    console.print(Panel(result or "[dim]No commits[/]", title="[bold cyan]Git Log[/]", border_style="cyan"))
    console.print()


def _show_history():
    if _HAS_AI:
        chats = get_all_chats()
        if chats:
            table = Table(box=ROUNDED, title="[bold cyan]Chat History[/]", title_style="bold cyan", border_style="cyan", padding=(0, 1))
            table.add_column("Session", style="dim", min_width=16)
            table.add_column("Last Message", style="white")
            for c in chats[:15]:
                table.add_row(c["chat_id"][:16], c["last_message"][:50])
            console.print()
            console.print(table)
        else:
            console.print("[dim]No chat history yet.[/]")
    else:
        console.print("[dim]AI module not available.[/]")


# ============================================================================
# INTERACTIVE SELECTIONS
# ============================================================================

def _select_agent():
    options = []
    for a in AGENTS:
        marker = "  [green]◄[/]" if a["id"] == _state["agent"]["id"] else ""
        options.append({
            "display": f"{a['icon']}  {a['name']}  [dim]- {a['desc']}[/]{marker}",
            "value": a,
        })
    result = _fuzzy_select("Select Agent:", options, key_func=lambda o: o.get("name", ""))
    if result:
        _state["agent"] = result
        console.print(f"\n  [green]✓[/] Agent: [bold]{result['name']}[/]  [dim]{result['desc']}[/]\n")
    return result


def _select_model():
    options = []
    for m in MODELS:
        free = " [green]✓free[/]" if m["free"] else " [red]paid[/]"
        marker = "  [green]◄[/]" if m["id"] == _state["model"]["id"] else ""
        options.append({
            "display": f"[{m['color']}]{m['name']}[/]  [dim]{m['provider']}[/]{free}{marker}",
            "value": m,
        })
    result = _fuzzy_select("Select Model:", options, key_func=lambda o: o.get("name", ""))
    if result:
        _state["model"] = result
        console.print(f"\n  [green]✓[/] Model: [bold]{result['name']}[/]  [dim]{result['provider']}[/]\n")
    return result


def _select_plugin():
    options = []
    for p in PLUGINS:
        options.append({
            "display": f"{p['icon']}  {p['name']}  [dim]- {p['desc']}[/]",
            "value": p,
        })
    result = _fuzzy_select("Select Plugin:", options, key_func=lambda o: o.get("name", ""))
    if result:
        action_options = [{"display": f"  {a}", "value": a} for a in result["actions"]]
        action = _fuzzy_select(f"Select {result['name']} action:", action_options, key_func=lambda o: o)
        if action:
            return result, action
    return None, None


def _select_session():
    sessions = _get_sessions()
    if not sessions:
        console.print("[dim]No sessions found. Type /new to create one.[/]")
        return None
    options = []
    for s in sessions:
        msg_count = len(s.get("messages", []))
        options.append({
            "display": f"💬  {s.get('name', s['id'][:8])}  [dim]{msg_count} messages  {s.get('agent', 'general')}[/]",
            "value": s,
        })
    return _fuzzy_select("Select Session:", options, key_func=lambda o: o.get("name", ""))


def _show_commit():
    result = _run_cmd("git status --short")
    if not result.strip():
        console.print("[dim]No changes to commit.[/]")
        return
    console.print(Panel(result[:2000], title="[bold cyan]Changes[/]", border_style="cyan"))
    try:
        msg = input("  Commit message: ").strip()
    except (EOFError, KeyboardInterrupt):
        console.print("\n[dim]Cancelled[/]")
        return
    if msg:
        r = _run_cmd(f'git add -A && git commit -m "{msg}"')
        console.print(f"[green]✓[/] {r[:500]}")


# ============================================================================
# COMMAND HANDLER
# ============================================================================

def _handle_slash(inp):
    if not inp.startswith("/"):
        return False
    parts = inp[1:].split(" ", 1)
    cmd = parts[0].lower().strip()
    val = parts[1].strip() if len(parts) > 1 else ""

    if cmd in {"help", "?"}:
        _show_help()
        return True

    if cmd == "clear":
        console.clear()
        return True

    if cmd == "connect":
        _show_connect()
        return True

    if cmd == "agents":
        if not val and _HAS_Q:
            _select_agent()
        else:
            _show_agents()
        return True

    if cmd == "model":
        if not val and _HAS_Q:
            _select_model()
        else:
            _show_models()
        return True

    if cmd == "skills":
        _show_skills()
        return True

    if cmd == "plugins":
        if not val and _HAS_Q:
            plugin, action = _select_plugin()
            if plugin and action:
                console.print(f"  [cyan]▶[/] {plugin['name']} > {action}")
        else:
            _show_plugins()
        return True

    if cmd == "init":
        _show_init()
        return True

    if cmd == "config":
        _show_config()
        return True

    if cmd == "doctor":
        _show_doctor()
        return True

    if cmd == "device":
        _show_device()
        return True

    if cmd == "history":
        _show_history()
        return True

    if cmd == "palette":
        result = _command_palette()
        if result:
            return _handle_slash(result)
        return True

    # Session commands
    if cmd == "new":
        chat_id = f"cli_{uuid4().hex[:8]}"
        if _HAS_AI:
            chat_id = f"cli_{get_device_id()}_{uuid4().hex[:8]}"
        _state["chat_id"] = chat_id
        _state["history"] = []
        _state["session_name"] = val or f"Session {datetime.now().strftime('%H:%M')}"
        console.print(f"\n  [green]✓[/] New session: [bold]{_state['session_name']}[/]  [dim]{chat_id[:12]}[/]\n")
        return True

    if cmd == "session":
        if _HAS_Q:
            sel = _select_session()
            if sel:
                _state["chat_id"] = sel["id"]
                _state["history"] = sel.get("messages", [])
                _state["session_name"] = sel.get("name", "")
                console.print(f"\n  [green]✓[/] Switched to: [bold]{sel.get('name', sel['id'][:12])}[/]\n")
        else:
            _show_sessions()
        return True

    if cmd == "sessions":
        _show_sessions()
        return True

    if cmd == "resume":
        sessions = _get_sessions()
        if sessions:
            sel = sessions[0]
            _state["chat_id"] = sel["id"]
            _state["history"] = sel.get("messages", [])
            _state["session_name"] = sel.get("name", "")
            console.print(f"\n  [green]✓[/] Resumed: [bold]{sel.get('name', sel['id'][:12])}[/]\n")
        else:
            console.print("[dim]No sessions to resume.[/]")
        return True

    # File commands
    if cmd in {"files", "ls"}:
        console.print()
        content = _list_files(val or ".")
        console.print(Panel(content or "[dim]No files[/]", title="[bold cyan]Files[/]", border_style="cyan"))
        console.print()
        return True

    if cmd == "read":
        if not val:
            console.print("[dim]Usage: /read <path>[/]")
        else:
            content = _read_file(val)
            console.print()
            try:
                syntax = Syntax(content, val.split(".")[-1] if "." in val else "text", theme="monokai", line_numbers=True)
                console.print(Panel(syntax, title=f"[bold cyan]{val}[/]", border_style="cyan"))
            except Exception:
                console.print(Panel(content, title=f"[bold cyan]{val}[/]", border_style="cyan"))
            console.print()
        return True

    if cmd in {"open", "cat"}:
        if not val:
            console.print("[dim]Usage: /open <any-file-path>[/]")
            console.print("[dim]Examples:[/]")
            console.print("[dim]  /open C:\\Users\\ADMIN\\Downloads\\data.xlsx[/]")
            console.print("[dim]  /open D:\\projects\\README.md[/]")
            console.print("[dim]  /open C:\\Users\\ADMIN\\Downloads\\archive.zip[/]")
        else:
            console.print()
            content = _read_file(val)
            console.print(Panel(content, title=f"[bold cyan]{val}[/]", border_style="cyan"))
            console.print()
        return True

    if cmd in {"ls", "dir"}:
        if not val:
            val = "."
        console.print()
        content = _list_files(val)
        console.print(Panel(content or "[dim]No files[/]", title=f"[bold cyan]{val}[/]", border_style="cyan"))
        console.print()
        return True

    if cmd == "write":
        if not val:
            console.print("[dim]Usage: /write <path>  (then describe what to write in chat)[/]")
        else:
            console.print(f"[dim]Writing to {val} - describe content in your message[/]")
        return True

    if cmd == "edit":
        if not val:
            console.print("[dim]Usage: /edit <path>[/]")
        else:
            console.print(f"[dim]Editing {val} - describe changes in your message[/]")
        return True

    if cmd == "run":
        if not val:
            console.print("[dim]Usage: /run <command>[/]")
        else:
            console.print(f"\n  [cyan]$[/] [dim]{val}[/]")
            result = _run_cmd(val)
            console.print(Panel(result[:3000], border_style="dim"))
            console.print()
        return True

    # Git commands
    if cmd == "diff":
        _show_diff()
        return True

    if cmd == "log":
        _show_log()
        return True

    if cmd == "commit":
        _show_commit()
        return True

    if cmd == "git":
        if not val:
            console.print("[dim]Usage: /git <args>  e.g. /git status[/]")
        else:
            result = _run_cmd(f"git {val}")
            console.print(f"\n  [dim]git {val}[/]")
            console.print(Panel(result[:3000], border_style="dim"))
            console.print()
        return True

    # Tool commands
    if cmd == "web":
        if not val:
            console.print("[dim]Usage: /web <search query>[/]")
        else:
            console.print(f"\n  [cyan]Searching:[/] {val}")
            try:
                import urllib.request, urllib.parse
                url = f"https://www.google.com/search?q={urllib.parse.quote(val)}"
                result = _run_cmd(f'curl -sL "{url}" -H "User-Agent: Mozilla/5.0" 2>nul | head -c 2000')
                console.print(Panel(result[:1000] or "[dim]No results[/]", border_style="dim"))
            except Exception:
                console.print("[dim]Web search not available[/]")
            console.print()
        return True

    if cmd == "img":
        if not val:
            console.print("[dim]Usage: /img <image_path>[/]")
        else:
            p = _safe(val)
            if p.exists():
                sz = p.stat().st_size
                console.print(Panel(f"Image: [bold]{val}[/]\nSize: {sz} bytes\nType: {p.suffix}", title="[bold cyan]Image[/]", border_style="cyan"))
            else:
                console.print(f"[red]File not found: {val}[/]")
        return True

    if cmd == "compact":
        console.print(Panel("[cyan]Context compacted[/]  [dim]History truncated for next turn[/]", border_style="cyan"))
        return True

    if cmd == "cost":
        prov, model, has_key = _get_provider_info()
        table = Table(box=MINIMAL, border_style="cyan")
        table.add_column("Key", style="bold cyan")
        table.add_column("Value")
        table.add_row("Provider", prov)
        table.add_row("Model", model)
        table.add_row("Agent", f"{_state['agent']['icon']} {_state['agent']['name']}")
        table.add_row("Session", _state.get("session_name") or "none")
        table.add_row("API Key", "[green]configured[/]" if has_key else "[red]not set[/]")
        console.print()
        console.print(Panel(table, title="[bold cyan]Token Estimate[/]", border_style="cyan"))
        console.print()
        return True

    if cmd == "quit":
        # Save session before quit
        if _state.get("chat_id") and _state.get("history"):
            _save_session(_state["chat_id"], _state["history"], _state.get("session_name", ""))
        console.print("\n  [dim]bye[/]\n")
        raise SystemExit(0)

    console.print(f"[dim]Unknown command. Type /help or press Ctrl+P for command palette[/]")
    return True


def _show_sessions():
    tbl = _list_sessions()
    if tbl:
        console.print()
        console.print(tbl)
        console.print()
    else:
        console.print("[dim]No sessions. Type /new to create one.[/]")


# ============================================================================
# AI CORE
# ============================================================================

def _try_direct_command(inp):
    import subprocess
    lower = inp.lower().strip()

    git_cmds = {
        "git status": "git status",
        "git log": "git log --oneline -10",
        "git diff": "git diff",
        "git branches": "git branch -a",
        "git branch": "git branch",
        "git remote": "git remote -v",
        "git pull": "git pull",
        "git push": "git push",
        "git add": "git add .",
        "git commit": "git commit -m 'Update from AuraCode'",
        "connect to git": "git status && git remote -v",
        "check git": "git status",
        "show git": "git status && git log --oneline -5",
        "git stash": "git stash",
        "git checkout": "git checkout main",
        "git clone": None,
    }

    for pattern, cmd in git_cmds.items():
        if lower == pattern or lower.startswith(pattern + " "):
            if cmd is None:
                return f"Usage: git clone <url>\nExample: git clone https://github.com/user/repo.git"
            r = subprocess.run(cmd, shell=True, text=True, capture_output=True, timeout=30, cwd=str(WORKSPACE))
            output = (r.stdout + r.stderr).strip()
            if not output:
                output = f"Command executed: {cmd}"
            return f"$ {cmd}\n\n{output}"

    if lower.startswith("run ") or lower.startswith("execute "):
        cmd = inp[4:].strip() if lower.startswith("run ") else inp[8:].strip()
        r = subprocess.run(cmd, shell=True, text=True, capture_output=True, timeout=60, cwd=str(WORKSPACE))
        output = (r.stdout + r.stderr).strip()
        return f"$ {cmd}\n\n{output}" or f"$ {cmd}\n\nDone."

    if lower.startswith("dir ") or lower.startswith("ls "):
        path = inp[4:].strip() if lower.startswith("dir ") else inp[3:].strip()
        return _list_files(path or ".")

    if lower.startswith("open ") or lower.startswith("read "):
        path = inp[5:].strip() if lower.startswith("open ") else inp[5:].strip()
        return _read_file(path)

    if lower.startswith("create ") or lower.startswith("write "):
        parts = inp.split(" ", 1)
        if len(parts) > 1:
            return f"To create a file, use: /write <path>\nThen describe what to write."

    if lower in ("help me", "what can you do", "what can you"):
        return """I can:
- Run any command: git, python, dir, etc.
- Read any file: Excel, PDF, ZIP, code, text
- Write and edit files anywhere
- List files in any directory
- Connect to Git repositories
- Execute code in Python/JavaScript

Just tell me what to do!"""

    return None


def _ask(inp, tool_results="", history=None):
    if not _HAS_AI:
        return {"message": "AI module not available.", "actions": []}
    mem = build_memory_context()
    sys = SYSTEM_PROMPT
    agent_sys = _state["agent"].get("system", "")
    if agent_sys:
        sys += f"\n\nAGENT ROLE:\n{agent_sys}"
    if mem:
        sys += f"\n\nUSER MEMORY:\n{mem}"
    msgs = [{"role": "system", "content": sys}]
    if history:
        for m in history[-20:]:
            msgs.append({"role": m.get("role", "user"), "content": m.get("content", "")})
    if tool_results:
        msgs.append({"role": "user", "content": f"Tool results:\n{tool_results}\n\nUser: {inp}"})
    else:
        msgs.append({"role": "user", "content": inp})
    content = chat_completion(msgs, temperature=0.1, json_mode=True)
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        return {"message": content, "actions": []}


def _run_turn(inp, chat_id):
    if _HAS_AI:
        extract_facts_from_message(inp)
        store_chat_message(chat_id, "user", inp)
        history = get_chat_history(chat_id, limit=30)
    else:
        history = _state.get("history", [])

    direct = _try_direct_command(inp)
    if direct:
        console.print()
        console.print(Panel(direct, border_style="cyan", padding=(0, 1)))
        console.print()
        _state["history"].append({"role": "assistant", "content": direct})
        if _HAS_AI:
            store_chat_message(chat_id, "assistant", direct, _state["agent"]["id"])
        return

    tool_results = ""
    for _ in range(5):
        with console.status(f"[bold cyan]{_state['agent']['icon']} Thinking...[/]", spinner="dots"):
            try:
                resp = _ask(inp, tool_results, history)
            except Exception as e:
                err = str(e)
                if "400" in err or "401" in err or "403" in err or "Invalid" in err.lower():
                    console.print(f"\n  [red]✗ AI Error:[/] {err}")
                    console.print("  [yellow]Your API key may be invalid or missing.[/]")
                    console.print("  [dim]Fix: type /connect to set up a working AI provider[/]\n")
                elif "No AI backend" in err or "not available" in err.lower():
                    console.print(f"\n  [red]✗ No AI backend available[/]")
                    console.print("  [yellow]Fix: type /connect to set up an AI provider[/]")
                    console.print("  [dim]Free options: Google Gemini, Groq, DeepSeek[/]\n")
                else:
                    console.print(f"\n  [red]✗ error:[/] {err}\n")
                return

        msg = resp.get("message", "")
        actions = resp.get("actions", [])
        if msg:
            console.print()
            console.print(Panel(msg, border_style="cyan", padding=(0, 1)))
            console.print()
            _state["history"].append({"role": "assistant", "content": msg})
            if _HAS_AI:
                store_chat_message(chat_id, "assistant", msg, _state["agent"]["id"])
        if not actions:
            return

        for a in actions:
            tool = a.get("tool", "")
            det = a.get("path", a.get("command", ""))
            if tool == "write_file":
                det = a.get("path", "")
            if det and len(str(det)) > 60:
                det = str(det)[:57] + "..."
            console.print(f"  [magenta]▶[/] [cyan]{tool}[/] [dim]{det}[/]")

        with console.status("[bold cyan]Executing...[/]", spinner="dots"):
            tool_results = _exec(actions)

        if tool_results.strip():
            for line in tool_results.strip().split("\n")[:10]:
                console.print(f"  {line}")
            console.print()


# ============================================================================
# AUTO-SETUP WIZARD
# ============================================================================

def _check_ai_ready():
    """Check if any cloud provider API key is available."""
    from app.config import _read_env
    _read_env()
    for key in ["GOOGLE_API_KEY", "OPENAI_API_KEY", "GROQ_API_KEY", "DEEPSEEK_API_KEY", "ANTHROPIC_API_KEY"]:
        val = os.getenv(key, "").strip()
        if val and len(val) > 5:
            return True
    return False


def _auto_setup_wizard():
    """Interactive setup wizard - guides user to connect an AI provider."""
    console.print(Panel("[bold cyan]Quick AI Setup[/]\n[dim]Get free key: https://aistudio.google.com/app/apikey[/]", border_style="cyan"))

    options = [
        {
            "display": "G  Google Gemini  [green](free, recommended)[/]  [dim]aistudio.google.com[/]",
            "value": {"name": "Google Gemini", "id": "google", "key_env": "GOOGLE_API_KEY", "model_env": "GOOGLE_CHAT_MODEL", "model": "gemini-2.0-flash", "url": "https://aistudio.google.com/app/apikey", "steps": "1. Click link\n2. Sign in with Google\n3. Click 'Create API Key'\n4. Copy the key"},
        },
        {
            "display": "Q  Groq (Fastest)  [green](free)[/]  [dim]console.groq.com[/]",
            "value": {"name": "Groq", "id": "groq", "key_env": "GROQ_API_KEY", "model_env": "GROQ_CHAT_MODEL", "model": "llama-3.3-70b-versatile", "url": "https://console.groq.com/keys", "steps": "1. Click link\n2. Sign up free\n3. Create API key\n4. Copy the key"},
        },
        {
            "display": "D  DeepSeek  [green](free)[/]  [dim]platform.deepseek.com[/]",
            "value": {"name": "DeepSeek", "id": "deepseek", "key_env": "DEEPSEEK_API_KEY", "model_env": "DEEPSEEK_CHAT_MODEL", "model": "deepseek-chat", "url": "https://platform.deepseek.com/", "steps": "1. Click link\n2. Sign up\n3. Go to API Keys\n4. Create and copy key"},
        },
    ]

    console.print()
    result = _fuzzy_select("Choose an AI provider:", options, key_func=lambda o: o.get("name", ""))
    if not result:
        console.print("[dim]Skipped. Type /connect later to set up.[/]\n")
        return

    provider = result
        console.print()
    console.print(f"[bold cyan]{provider['name']}[/]")
    console.print(Panel(provider["steps"], border_style="cyan", title="[dim]Steps[/]"))
    console.print(f"[dim]Link: {provider['url']}[/]")
    console.print()

    try:
        api_key = input("  Paste your API key here: ").strip()
    except (EOFError, KeyboardInterrupt):
        console.print("\n[dim]Skipped. Type /connect later.[/]\n")
        return

    if not api_key or len(api_key) < 10:
        console.print("[yellow]Key too short. Type /connect later.[/]\n")
        return

    _set_env_var(provider["key_env"], api_key)
    _set_env_var("AI_PROVIDER", provider["id"])
    if provider["model_env"]:
        _set_env_var(provider["model_env"], provider["model"])

    console.print(f"\n[green]✓ Connected to {provider['name']}![/]")
    console.print(f"[green]✓ Model: {provider['model']}[/]")
    console.print("[dim]Starting chat...[/]\n")


# ============================================================================
# MAIN
# ============================================================================

def _print_header():
    prov, model, has_key = _get_provider_info()
    agent = _state["agent"]
    model_info = _state["model"]
    session = _state.get("session_name") or _state.get("chat_id", "")[:12] if _state.get("chat_id") else "new"

    header = Table(box=ROUNDED, border_style="cyan", padding=(0, 1), show_header=False)
    header.add_column("Content", ratio=1)
    header.add_row("[bold cyan]AuraCode[/]  [dim]v2.0[/]")
    header.add_row("[dim]OpenCode-style terminal coding agent[/]")
    header.add_row(f"[dim]Agent:[/] {agent['icon']} [bold]{agent['name']}[/]  [dim]Model:[/] {model_info['name']}  [dim]Provider:[/] {prov}")
    header.add_row(f"[dim]Session:[/] [cyan]{session}[/]  [dim]Workspace:[/] {WORKSPACE.name}")
    if not has_key:
        header.add_row("[yellow]![/] [dim]No API key - type /connect to set up (free Google key)[/]")

    console.print()
    console.print(header)
    console.print("[dim]Type a message, / for commands, Ctrl+P for command palette[/]")
    console.print()


def main():
    _load_ai()
    chat_id = f"cli_{uuid4().hex[:8]}"
    if _HAS_AI:
        chat_id = f"cli_{get_device_id()}_{uuid4().hex[:8]}"
    _state["chat_id"] = chat_id

    # Try to resume last session
    sessions = _get_sessions()
    if sessions:
        last = sessions[0]
        _state["session_name"] = last.get("name", last["id"][:8])

    _print_header()

    # Auto-setup: check if any AI provider is configured
    if not _check_ai_ready():
        console.print("[yellow]No AI provider configured![/]")
        console.print("[dim]Set GOOGLE_API_KEY or OPENAI_API_KEY in .env[/]\n")
        console.print("[cyan]Quick setup:[/]")
        console.print("[dim]  1. Get free key: https://aistudio.google.com/app/apikey[/]")
        console.print("[dim]  2. Add to .env: GOOGLE_API_KEY=your_key[/]")
        console.print("[dim]  3. Restart AuraCode[/]\n")

    while True:
        try:
            agent_icon = _state["agent"]["icon"]
            u = console.input(f"  [green]{agent_icon}[/] ").strip()
        except (EOFError, KeyboardInterrupt):
            console.print("\n  [dim]bye[/]\n")
            break
        if not u:
            continue
        if u.lower() in {"exit", "quit", "/quit"}:
            if _state.get("chat_id") and _state.get("history"):
                _save_session(_state["chat_id"], _state["history"], _state.get("session_name", ""))
            console.print("\n  [dim]bye[/]\n")
            break
        # Ctrl+P command palette
        if u == "\x10":  # Ctrl+P
            result = _command_palette()
            if result:
                try:
                    _handle_slash(result)
                except SystemExit:
                    console.print("\n  [dim]bye[/]\n")
                    break
            continue
        try:
            if _handle_slash(u):
                continue
            _state["history"].append({"role": "user", "content": u})
            _run_turn(u, chat_id)
        except SystemExit:
            if _state.get("chat_id") and _state.get("history"):
                _save_session(_state["chat_id"], _state["history"], _state.get("session_name", ""))
            console.print("\n  [dim]bye[/]\n")
            break
        except KeyboardInterrupt:
            console.print("\n  [dim]interrupted[/]\n")
            continue
        except Exception as e:
            console.print(f"\n  [red]✗ error:[/] {e}\n")


def _is_server_running():
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.connect(("127.0.0.1", 18765))
        s.close()
        return True
    except ConnectionRefusedError:
        return False


def _start_server(host="0.0.0.0", port=18765):
    import socket
    import threading

    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    sock.bind((host, port))
    sock.listen(5)

    def _handle(conn):
        try:
            data = b""
            while True:
                chunk = conn.recv(65536)
                if not chunk:
                    break
                data += chunk
                if len(data) > 100000:
                    break
            req = json.loads(data.decode("utf-8"))
            if req.get("type") == "chat":
                _load_ai()
                msg = req.get("message", "")
                chat_id = req.get("chat_id", f"cli_{uuid4().hex[:8]}")
                _state["chat_id"] = chat_id
                _state["history"].append({"role": "user", "content": msg})
                direct = _try_direct_command(msg)
                if direct:
                    result = direct
                else:
                    from app.device import store_chat_message, get_chat_history
                    store_chat_message(chat_id, "user", msg)
                    history = get_chat_history(chat_id, limit=30)
                    result = _run_turn(msg, chat_id)
                conn.send(json.dumps({"type": "response", "message": result or ""}).encode("utf-8"))
            elif req.get("type") == "status":
                conn.send(json.dumps({"type": "status", "running": True}).encode("utf-8"))
            elif req.get("type") == "update":
                conn.send(json.dumps({"type": "update", "version": "2.0"}).encode("utf-8"))
        except Exception as e:
            try:
                conn.send(json.dumps({"type": "error", "message": str(e)}).encode("utf-8"))
            except Exception:
                pass
        finally:
            try:
                conn.close()
            except Exception:
                pass

    def _accept():
        while True:
            try:
                conn, addr = sock.accept()
                threading.Thread(target=_handle, args=(conn,), daemon=True).start()
            except Exception:
                pass

    threading.Thread(target=_accept, daemon=True).start()
    return sock


def _client_send(msg, chat_id=None):
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.connect(("127.0.0.1", 18765))
        s.send(json.dumps({"type": "chat", "message": msg, "chat_id": chat_id}).encode("utf-8"))
        data = b""
        while True:
            chunk = s.recv(65536)
            if not chunk:
                break
            data += chunk
            if len(data) > 100000:
                break
        return json.loads(data.decode("utf-8"))
    except ConnectionRefusedError:
        return None
    finally:
        try:
            s.close()
        except Exception:
            pass


def _auto_update():
    import subprocess
    try:
        r = subprocess.run("git pull", shell=True, text=True, capture_output=True, timeout=30, cwd=str(WORKSPACE))
        if "Already up to date" not in r.stdout:
            console.print("[green]Updated to latest version![/]")
            return True
    except Exception:
        pass
    return False


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--server":
        _load_ai()
        sock = _start_server()
        console.print("[green]AuraCode server started on port 18765[/]")
        console.print("[dim]Other devices can connect now[/]")
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            sock.close()
    elif len(sys.argv) > 1 and sys.argv[1] == "--update":
        _auto_update()
    elif _is_server_running():
        _load_ai()
        console.print("[dim]Connecting to running AuraCode...[/]")
        _state["chat_id"] = f"cli_{uuid4().hex[:8]}"
        while True:
            try:
                agent_icon = _state["agent"]["icon"]
                u = console.input(f"  [green]{agent_icon}[/] ").strip()
            except (EOFError, KeyboardInterrupt):
                console.print("\n  [dim]bye[/]\n")
                break
            if not u:
                continue
            if u.lower() in {"exit", "quit"}:
                break
            if u.startswith("/"):
                if _handle_slash(u):
                    continue
            resp = _client_send(u, _state["chat_id"])
            if resp and resp.get("type") == "response":
                msg = resp.get("message", "")
                if msg:
                    console.print()
                    console.print(Panel(msg, border_style="cyan", padding=(0, 1)))
                    console.print()
            elif resp:
                console.print(f"\n[yellow]{resp.get('message', 'Unknown error')}[/]\n")
            else:
                console.print("\n[yellow]Server not responding. Restart with: python auracode.py --server[/]\n")
                break
    else:
        _load_ai()
        console.print("[dim]Starting AuraCode server...[/]")
        _start_server()
        main()
